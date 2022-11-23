import os.path

from openpyxl import load_workbook as lw

way = '/home/alexander/Python_projects/test.xlsx'


def open_xl(way, *args):
    data = ''
    for sheet in args:
        wb = lw(way)
        ws = wb[f'{sheet}']
        for cell in range(9, 15 + 1):
            temp = ws[f'B{cell}'].value

            if temp == None:
                pass
            elif temp[1] == '.' or temp[2] == '.':
                data += '\n' + ws[f'B{cell}'].value
            else:
                data += ws[f'B{cell}'].value

    return data


def data_to_txt(way, *args):
    os.system(r' >/home/alexander/Python_projects/script.txt')
    my_file = open('/home/alexander/Python_projects/script.txt', 'w+')
    my_file.write(open_xl(way, *args))
    my_file.close()


if __name__ == '__main__':
    way = input('Введите путь до файла')
    name_file = input('Введите название файла')
    sheet_start = input('Введите с какого листа начинать')
    sheet_end = input('Введите до какого листа выполнять чтение')
    way = way + '/' + name_file
    data_to_txt(way, sheet_start, sheet_end)
